"""
Agent 知识库后端 - v3（可安全推 GitHub）
FastAPI + SQLite + 通义千问AI解析 + 飞书多维表格同步

运行方式: 
  1. 复制 .env.example 为 .env，填入你的密钥
  2. pip install fastapi uvicorn python-multipart PyPDF2 python-docx openpyxl httpx python-dotenv
  3. python main.py

访问: http://localhost:8000
"""
import os
import sys
import json
import sqlite3
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import uvicorn
import secrets

# 加载 .env 文件
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # 没装 python-dotenv 也能跑（手动设环境变量）

# ===================== 配置（从环境变量读取） =====================
BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
DB_PATH = BASE_DIR / "knowledge.db"
UPLOAD_DIR.mkdir(exist_ok=True)

# AI 模型（阿里云百炼/DashScope）
DASHSCOPE_API_KEY = os.environ.get("DASHSCOPE_API_KEY", "")
DASHSCOPE_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DASHSCOPE_MODEL = "qwen-plus"

# 飞书多维表格
FEISHU_APP_ID = os.environ.get("FEISHU_APP_ID", "")
FEISHU_APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "")
FEISHU_APP_TOKEN = os.environ.get("FEISHU_APP_TOKEN", "")
FEISHU_TABLE_ID = os.environ.get("FEISHU_TABLE_ID", "")

# 管理后台密码保护
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "juzi2024")

security = HTTPBasic()

def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    """验证管理后台密码"""
    correct_user = secrets.compare_digest(credentials.username, ADMIN_USERNAME)
    correct_pass = secrets.compare_digest(credentials.password, ADMIN_PASSWORD)
    if not (correct_user and correct_pass):
        raise HTTPException(status_code=401, detail="未授权", headers={"WWW-Authenticate": "Basic"})
    return credentials.username

app = FastAPI(title="Agent知识库后端", version="3.0.0", docs_url=None, redoc_url=None)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===================== 数据库初始化 =====================
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS submissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company TEXT NOT NULL, industry TEXT NOT NULL,
        scenario TEXT, extra TEXT, created_at TEXT DEFAULT (datetime('now', 'localtime')))""")
    c.execute("""CREATE TABLE IF NOT EXISTS files (
        id INTEGER PRIMARY KEY AUTOINCREMENT, submission_id INTEGER, category TEXT,
        original_name TEXT, saved_path TEXT, file_type TEXT, file_size INTEGER,
        created_at TEXT DEFAULT (datetime('now', 'localtime')),
        FOREIGN KEY (submission_id) REFERENCES submissions(id))""")
    c.execute("""CREATE TABLE IF NOT EXISTS knowledge_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT, submission_id INTEGER, file_id INTEGER,
        entry_type TEXT, title TEXT, content TEXT, metadata TEXT,
        created_at TEXT DEFAULT (datetime('now', 'localtime')),
        FOREIGN KEY (submission_id) REFERENCES submissions(id),
        FOREIGN KEY (file_id) REFERENCES files(id))""")
    conn.commit()
    conn.close()

init_db()

# ===================== 文件解析器 =====================
def parse_pdf(filepath: str) -> list[dict]:
    entries = []
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                entries.append({"type": "document", "title": f"第{i+1}页", "content": text.strip()})
    except Exception as e:
        entries.append({"type": "error", "title": "PDF解析失败", "content": str(e)})
    return entries

def parse_docx(filepath: str) -> list[dict]:
    entries = []
    try:
        from docx import Document
        doc = Document(filepath)
        current_title, current_content = "正文", []
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text: continue
            style_name = para.style.name if para.style else ""
            if style_name.startswith("Heading"):
                if current_content:
                    entries.append({"type": "document", "title": current_title, "content": "\n".join(current_content)})
                current_title, current_content = text, []
            else:
                current_content.append(text)
        if current_content:
            entries.append({"type": "document", "title": current_title, "content": "\n".join(current_content)})
    except Exception as e:
        entries.append({"type": "error", "title": "Word解析失败", "content": str(e)})
    return entries

def parse_excel(filepath: str) -> list[dict]:
    entries = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(c) if c is not None else "" for c in row)
                if row_str.strip().replace("|", "").strip():
                    rows_text.append(row_str)
            if rows_text:
                entries.append({"type": "table", "title": f"表格-{sheet_name}", "content": "\n".join(rows_text[:200])})
    except Exception as e:
        entries.append({"type": "error", "title": "Excel解析失败", "content": str(e)})
    return entries

def parse_txt(filepath: str) -> list[dict]:
    entries = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()
        qa_pattern = r'[问Q][：:]\s*(.+?)\s*[答A][：:]\s*(.+?)(?=[问Q][：:]|\Z)'
        qa_matches = re.findall(qa_pattern, text, re.DOTALL)
        if qa_matches:
            for q, a in qa_matches:
                entries.append({"type": "qa", "title": q.strip()[:100], "content": a.strip()[:500]})
        else:
            paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
            if not paragraphs:
                paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
            for i, para in enumerate(paragraphs[:50]):
                entries.append({"type": "document", "title": f"段落{i+1}", "content": para[:1000]})
    except Exception as e:
        entries.append({"type": "error", "title": "文本解析失败", "content": str(e)})
    return entries

def parse_image(filepath: str) -> list[dict]:
    import base64, httpx
    filename = Path(filepath).name
    if not DASHSCOPE_API_KEY:
        return [{"type": "image", "title": f"图片-{filename}", "content": f"[图片] {filename} - 未配置AI，需人工查看"}]
    try:
        with open(filepath, "rb") as f:
            img_data = base64.b64encode(f.read()).decode("utf-8")
        ext = Path(filepath).suffix.lower()
        mime = {".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",".gif":"image/gif",".bmp":"image/bmp"}.get(ext, "image/png")
        resp = httpx.post(f"{DASHSCOPE_BASE_URL}/chat/completions",
            headers={"Authorization": f"Bearer {DASHSCOPE_API_KEY}", "Content-Type": "application/json"},
            json={"model": "qwen-vl-plus", "messages": [{"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_data}"}},
                {"type": "text", "text": "请详细描述这张图片的内容。如果包含文字请提取所有文字。返回纯文本。"}
            ]}], "temperature": 0.3}, timeout=60.0)
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"]
            if content and content.strip():
                print(f"  ✅ 图片AI识别成功: {filename}")
                return [{"type": "document", "title": f"图片内容-{filename}", "content": content.strip()}]
    except Exception as e:
        print(f"  ⚠️ 图片AI识别异常: {filename}, {e}")
    return [{"type": "image", "title": f"图片-{filename}", "content": f"[图片] {filename} - AI识别失败，需人工查看"}]

def parse_file(filepath: str, filename: str) -> list[dict]:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf": return parse_pdf(filepath)
    elif ext in (".doc", ".docx"): return parse_docx(filepath)
    elif ext in (".xls", ".xlsx"): return parse_excel(filepath)
    elif ext in (".txt", ".csv", ".md"): return parse_txt(filepath)
    elif ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp"): return parse_image(filepath)
    elif ext in (".zip", ".rar"): return [{"type": "archive", "title": f"压缩包-{filename}", "content": "[压缩包] 需解压后处理"}]
    else: return [{"type": "unknown", "title": f"未知格式-{filename}", "content": f"不支持: {ext}"}]

# ===================== AI 智能解析 =====================
def ai_parse_text(text: str, filename: str = "", category: str = "") -> list[dict]:
    if not text or not text.strip() or not DASHSCOPE_API_KEY:
        return []
    max_chars = 6000
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...(文本过长已截断)"
    prompt = f"""你是一个知识库整理助手。请将以下内容整理为结构化的知识库条目。
要求：1.FAQ提取为type="qa" 2.流程/SOP提取为type="sop" 3.其他按主题分段type="document" 4.每条content不超过500字 5.返回纯JSON数组
文件名：{filename} 材料分类：{category}
内容：
{text}
返回JSON数组：[{{"type":"qa","title":"问题","content":"答案"}}]"""
    try:
        import httpx
        resp = httpx.post(f"{DASHSCOPE_BASE_URL}/chat/completions",
            headers={"Authorization": f"Bearer {DASHSCOPE_API_KEY}", "Content-Type": "application/json"},
            json={"model": DASHSCOPE_MODEL, "messages": [{"role": "user", "content": prompt}], "temperature": 0.3},
            timeout=60.0)
        if resp.status_code != 200: return []
        content = resp.json()["choices"][0]["message"]["content"].strip()
        if content.startswith("```"):
            content = re.sub(r'^```\w*\n?', '', content)
            content = re.sub(r'\n?```$', '', content)
        entries = json.loads(content)
        if isinstance(entries, list):
            return [{"type": e["type"], "title": str(e["title"])[:200], "content": str(e["content"])[:2000]}
                    for e in entries if isinstance(e, dict) and all(k in e for k in ("type","title","content"))]
    except Exception as e:
        print(f"AI解析异常: {e}")
    return []

def parse_file_with_ai(filepath: str, filename: str, category: str = "") -> list[dict]:
    rule_entries = parse_file(filepath, filename)
    if rule_entries and rule_entries[0]["type"] in ("image", "archive", "unknown", "error"):
        return rule_entries
    if rule_entries and any("图片内容-" in e.get("title","") for e in rule_entries):
        return rule_entries
    all_text = "\n\n".join(
        f"{e['title']}:\n{e['content']}" if e['title'] != f"段落{i+1}" else e['content']
        for i, e in enumerate(rule_entries))
    if not all_text.strip(): return rule_entries
    ai_entries = ai_parse_text(all_text, filename, category)
    if ai_entries:
        print(f"  ✅ AI解析成功: {filename} → {len(ai_entries)}条")
        return ai_entries
    print(f"  ⚠️ AI解析失败，回退规则解析: {filename} → {len(rule_entries)}条")
    return rule_entries

# ===================== 飞书同步 =====================
def get_feishu_token():
    if not FEISHU_APP_ID or not FEISHU_APP_SECRET:
        return None
    import requests as req
    resp = req.post("https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET})
    data = resp.json()
    return data["tenant_access_token"] if data.get("code") == 0 else None

def upload_file_to_feishu(token: str, filepath: str, filename: str) -> str:
    import requests as req
    with open(filepath, "rb") as f:
        resp = req.post("https://open.feishu.cn/open-apis/drive/v1/medias/upload_all",
            headers={"Authorization": f"Bearer {token}"},
            data={"file_name": filename, "parent_type": "bitable_file", "parent_node": FEISHU_APP_TOKEN, "size": str(os.path.getsize(filepath))},
            files={"file": (filename, f, "application/octet-stream")})
    data = resp.json()
    return data["data"]["file_token"] if data.get("code") == 0 else ""

def sync_to_feishu(company, industry, file_count, scenario="", uploaded_files=None):
    if not FEISHU_APP_TOKEN or not FEISHU_TABLE_ID:
        print("  ⚠️ 飞书未配置，跳过同步")
        return False
    import requests as req
    token = get_feishu_token()
    if not token: return False
    file_tokens, categories = [], []
    if uploaded_files:
        for f_info in uploaded_files:
            fp, fn, cat = f_info.get("path",""), f_info.get("name",""), f_info.get("category","未分类")
            if fp and os.path.exists(fp):
                ft = upload_file_to_feishu(token, fp, fn)
                if ft:
                    file_tokens.append({"file_token": ft})
                    categories.append(f"{cat}: {fn}")
    record = {"fields": {"企业名称": company, "行业": industry, "提交日期": int(datetime.now().timestamp()*1000), "文件数": file_count}}
    if scenario and scenario.strip(): record["fields"]["场景需求"] = scenario.strip()
    if file_tokens: record["fields"]["附件"] = file_tokens
    if categories: record["fields"]["附件分类"] = "\n".join(categories)
    resp = req.post(f"https://open.feishu.cn/open-apis/bitable/v1/apps/{FEISHU_APP_TOKEN}/tables/{FEISHU_TABLE_ID}/records",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"}, json=record)
    ok = resp.json().get("code") == 0
    print(f"  {'✅' if ok else '⚠️'} 飞书同步{'成功' if ok else '失败'}: {company}")
    return ok

# ===================== API 路由 =====================
@app.post("/api/submit")
async def submit_form(company: str = Form(...), industry: str = Form(...), scenario: str = Form(""),
                      extra: str = Form(""), files: list[UploadFile] = File(default=[]), categories: str = Form("")):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO submissions (company, industry, scenario, extra) VALUES (?, ?, ?, ?)", (company, industry, scenario, extra))
    submission_id = c.lastrowid
    try: cat_map = json.loads(categories) if categories else {}
    except: cat_map = {}
    company_dir = UPLOAD_DIR / company.replace(" ","_").replace("/","_")
    company_dir.mkdir(parents=True, exist_ok=True)
    results = []
    for file in files:
        if not file.filename: continue
        category = cat_map.get(file.filename, "未分类")
        cat_dir = company_dir / category.replace(" ","_").replace("/","_")
        cat_dir.mkdir(parents=True, exist_ok=True)
        safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        save_path = cat_dir / safe_name
        content = await file.read()
        with open(save_path, "wb") as f: f.write(content)
        c.execute("INSERT INTO files (submission_id, category, original_name, saved_path, file_type, file_size) VALUES (?,?,?,?,?,?)",
                  (submission_id, category, file.filename, str(save_path), file.content_type, len(content)))
        file_id = c.lastrowid
        entries = parse_file_with_ai(str(save_path), file.filename, category)
        for entry in entries:
            c.execute("INSERT INTO knowledge_entries (submission_id, file_id, entry_type, title, content, metadata) VALUES (?,?,?,?,?,?)",
                      (submission_id, file_id, entry["type"], entry["title"], entry["content"],
                       json.dumps({"source_file": file.filename, "category": category}, ensure_ascii=False)))
        results.append({"filename": file.filename, "category": category, "entries_count": len(entries)})
    conn.commit(); conn.close()
    try:
        feishu_files = []
        conn2 = sqlite3.connect(DB_PATH); c2 = conn2.cursor()
        for r in results:
            c2.execute("SELECT saved_path, original_name, category FROM files WHERE submission_id=? AND original_name=?", (submission_id, r["filename"]))
            row = c2.fetchone()
            if row: feishu_files.append({"path": row[0], "name": row[1], "category": row[2]})
        conn2.close()
        sync_to_feishu(company, industry, len(results), scenario, feishu_files)
    except Exception as e:
        print(f"  ⚠️ 飞书同步异常: {e}")
    return {"success": True, "submission_id": submission_id, "company": company, "files_processed": len(results), "details": results}

@app.get("/api/submissions")
async def list_submissions(user: str = Depends(verify_admin)):
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row; c = conn.cursor()
    c.execute("SELECT company, industry, COUNT(*) as submit_count, MAX(created_at) as last_submit, GROUP_CONCAT(id) as submission_ids FROM submissions GROUP BY company ORDER BY last_submit DESC")
    rows = [dict(r) for r in c.fetchall()]
    for row in rows:
        ids = row["submission_ids"].split(","); ph = ",".join("?"*len(ids))
        c.execute(f"SELECT COUNT(*) FROM files WHERE submission_id IN ({ph})", ids); row["file_count"] = c.fetchone()[0]
        c.execute(f"SELECT COUNT(*) FROM knowledge_entries WHERE submission_id IN ({ph})", ids); row["entry_count"] = c.fetchone()[0]
    conn.close()
    return {"submissions": rows}

@app.get("/api/knowledge/{submission_id}")
async def get_knowledge(submission_id: str, user: str = Depends(verify_admin)):
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row; c = conn.cursor()
    ids = [s.strip() for s in submission_id.split(",")]; ph = ",".join("?"*len(ids))
    c.execute(f"SELECT * FROM knowledge_entries WHERE submission_id IN ({ph}) ORDER BY entry_type, id", ids)
    entries = [dict(r) for r in c.fetchall()]
    c.execute(f"SELECT * FROM files WHERE submission_id IN ({ph}) ORDER BY category, id", ids)
    files = [dict(r) for r in c.fetchall()]
    conn.close()
    grouped = {}
    for e in entries:
        grouped.setdefault(e["entry_type"], []).append(e)
    return {"submission_id": submission_id, "total_entries": len(entries), "files": files, "grouped": grouped}

@app.get("/api/file/{file_id}")
async def download_file(file_id: int, user: str = Depends(verify_admin)):
    from fastapi.responses import FileResponse
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row; c = conn.cursor()
    c.execute("SELECT * FROM files WHERE id = ?", (file_id,)); row = c.fetchone(); conn.close()
    if not row: raise HTTPException(404, "文件不存在")
    fp = Path(row["saved_path"])
    if not fp.exists(): raise HTTPException(404, "文件已删除")
    return FileResponse(str(fp), filename=row["original_name"], media_type=row["file_type"] or "application/octet-stream")

@app.get("/api/stats")
async def get_stats(user: str = Depends(verify_admin)):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM submissions"); ts = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM files"); tf = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM knowledge_entries"); te = c.fetchone()[0]
    c.execute("SELECT entry_type, COUNT(*) FROM knowledge_entries GROUP BY entry_type")
    td = {r[0]: r[1] for r in c.fetchall()}; conn.close()
    return {"total_submissions": ts, "total_files": tf, "total_knowledge_entries": te, "entry_type_distribution": td}

# ===================== 页面路由 =====================
FRONTEND_DIR = BASE_DIR.parent

@app.get("/form", response_class=HTMLResponse)
async def serve_form():
    p = FRONTEND_DIR / "index2.html"
    return HTMLResponse(p.read_text(encoding="utf-8")) if p.exists() else HTMLResponse("<h1>Not found</h1>", 404)

@app.get("/index2.html", response_class=HTMLResponse)
async def serve_form_alias():
    p = FRONTEND_DIR / "index2.html"
    return HTMLResponse(p.read_text(encoding="utf-8")) if p.exists() else HTMLResponse("<h1>Not found</h1>", 404)

@app.get("/", response_class=HTMLResponse)
async def serve_landing():
    p = FRONTEND_DIR / "landing.html"
    return HTMLResponse(p.read_text(encoding="utf-8")) if p.exists() else HTMLResponse("<h1>Not found</h1>", 404)

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(user: str = Depends(verify_admin)):
    """管理后台（需要密码）"""
    admin_html = """<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>知识库管理后台</title>
<style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;background:#f5f6fa;padding:24px}.container{max-width:1000px;margin:0 auto}h1{font-size:24px;margin-bottom:20px;color:#1a1a2e}.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:28px}.stat-card{background:#fff;border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,0.06)}.stat-card .num{font-size:32px;font-weight:700;color:#667eea}.stat-card .label{font-size:14px;color:#666;margin-top:4px}.section{background:#fff;border-radius:12px;padding:24px;margin-bottom:20px;box-shadow:0 2px 8px rgba(0,0,0,0.06)}.section h2{font-size:18px;margin-bottom:16px}table{width:100%;border-collapse:collapse}th,td{text-align:left;padding:10px 12px;border-bottom:1px solid #eee;font-size:14px}th{font-weight:600;color:#555;background:#f8f9fb}.btn{padding:6px 14px;background:#667eea;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px}.btn:hover{background:#5a6fd6}.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:500}.badge-qa{background:#e8f4e8;color:#2d7a2d}.badge-doc{background:#e8f0ff;color:#2d5aa0}.entry-item{padding:10px;border-left:3px solid #667eea;margin-bottom:8px;background:#fff;border-radius:0 6px 6px 0}.entry-item .et{font-size:12px;color:#888}.entry-item .ec{font-size:13px;color:#333;margin-top:4px;white-space:pre-wrap;max-height:400px;overflow-y:auto}.empty{text-align:center;color:#999;padding:40px}</style></head>
<body><div class="container"><h1>📚 Agent 知识库管理后台</h1><div class="stats" id="stats"></div><div class="section"><h2>📋 企业提交记录</h2><table><thead><tr><th>企业</th><th>行业</th><th>提交次数</th><th>文件数</th><th>知识条目</th><th>最近提交</th><th>操作</th></tr></thead><tbody id="subBody"></tbody></table></div><div class="section" id="detailSection" style="display:none"><h2>📖 知识库内容 - <span id="detailCompany"></span></h2><div id="fileList"></div><div id="detailContent"></div></div></div>
<script>
const cred=btoa('""" + ADMIN_USERNAME + ":" + ADMIN_PASSWORD + """');
const H={headers:{'Authorization':'Basic '+cred}};
async function loadStats(){const r=await fetch('/api/stats',H);const d=await r.json();document.getElementById('stats').innerHTML=`<div class="stat-card"><div class="num">${d.total_submissions}</div><div class="label">企业提交</div></div><div class="stat-card"><div class="num">${d.total_files}</div><div class="label">文件上传</div></div><div class="stat-card"><div class="num">${d.total_knowledge_entries}</div><div class="label">知识条目</div></div><div class="stat-card"><div class="num">${Object.keys(d.entry_type_distribution).length}</div><div class="label">内容类型</div></div>`}
async function loadSubs(){const r=await fetch('/api/submissions',H);const d=await r.json();const tb=document.getElementById('subBody');if(!d.submissions.length){tb.innerHTML='<tr><td colspan="7" class="empty">暂无记录</td></tr>';return}tb.innerHTML=d.submissions.map(s=>`<tr><td><b>${s.company}</b></td><td>${s.industry}</td><td>${s.submit_count}次</td><td>${s.file_count}个</td><td>${s.entry_count}条</td><td>${s.last_submit}</td><td><button class="btn" onclick="showD('${s.submission_ids}','${s.company}')">查看</button></td></tr>`).join('')}
async function showD(ids,co){document.getElementById('detailSection').style.display='block';document.getElementById('detailCompany').textContent=co;const r=await fetch('/api/knowledge/'+ids,H);const d=await r.json();const fl=document.getElementById('fileList'),ct=document.getElementById('detailContent');if(d.files&&d.files.length){let h='<h3>📂 文件清单</h3><table><thead><tr><th>分类</th><th>文件名</th><th>操作</th></tr></thead><tbody>';d.files.forEach(f=>{h+=`<tr><td>${f.category}</td><td>${f.original_name}</td><td><a href="/api/file/${f.id}" style="color:#667eea">下载</a></td></tr>`});h+='</tbody></table>';fl.innerHTML=h}else{fl.innerHTML=''}if(!d.total_entries){ct.innerHTML='<div class="empty">暂无条目</div>';return}let html='';const labels={qa:'问答对',document:'文档段落',table:'表格',sop:'SOP',image:'图片'};for(const[t,entries]of Object.entries(d.grouped)){html+=`<h3 style="margin:16px 0 8px"><span class="badge badge-${t=='qa'?'qa':'doc'}">${labels[t]||t}</span> (${entries.length}条)</h3>`;entries.slice(0,20).forEach(e=>{html+=`<div class="entry-item"><div class="et">${e.title}</div><div class="ec">${(e.content||'').substring(0,300).replace(/</g,'&lt;')}</div></div>`})}ct.innerHTML=html;document.getElementById('detailSection').scrollIntoView({behavior:'smooth'})}
loadStats();loadSubs()
</script></body></html>"""
    return HTMLResponse(admin_html)

# ===================== 启动 =====================
if __name__ == "__main__":
    print("=" * 50)
    print("  Agent 知识库后端 v3")
    print("=" * 50)
    print(f"  表单: http://localhost:8000/form")
    print(f"  后台: http://localhost:8000/admin")
    print(f"  API:  http://localhost:8000/docs")
    missing = []
    if not DASHSCOPE_API_KEY: missing.append("DASHSCOPE_API_KEY")
    if not FEISHU_APP_ID: missing.append("FEISHU_APP_ID")
    if missing: print(f"  ⚠️  未配置: {', '.join(missing)}（部分功能不可用）")
    print("=" * 50)
    uvicorn.run(app, host="0.0.0.0", port=8000)
