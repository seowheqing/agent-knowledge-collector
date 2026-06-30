"""
文件解析器模块 - 支持 PDF/Word/Excel/TXT/图片
"""
import re
import base64
from pathlib import Path

import httpx

from config import DASHSCOPE_API_KEY, DASHSCOPE_BASE_URL
from utils import retry


def parse_pdf(filepath: str) -> list[dict]:
    entries = []
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                entries.append({"type": "document", "title": f"第{i+1}页", "content": text.strip()})
    except Exception as e:
        entries.append({"type": "error", "title": "PDF解析失败", "content": str(e)})
    return entries


def parse_docx(filepath: str) -> list[dict]:
    entries = []
    try:
        from docx import Document
        doc = Document(filepath)
        current_title, current_content = "正文", []
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            style_name = para.style.name if para.style else ""
            if style_name.startswith("Heading"):
                if current_content:
                    entries.append({"type": "document", "title": current_title, "content": "\n".join(current_content)})
                current_title, current_content = text, []
            else:
                current_content.append(text)
        if current_content:
            entries.append({"type": "document", "title": current_title, "content": "\n".join(current_content)})
    except Exception as e:
        entries.append({"type": "error", "title": "Word解析失败", "content": str(e)})
    return entries


def parse_excel(filepath: str) -> list[dict]:
    entries = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(c) if c is not None else "" for c in row)
                if row_str.strip().replace("|", "").strip():
                    rows_text.append(row_str)
            if rows_text:
                entries.append({"type": "table", "title": f"表格-{sheet_name}", "content": "\n".join(rows_text[:200])})
    except Exception as e:
        entries.append({"type": "error", "title": "Excel解析失败", "content": str(e)})
    return entries


def parse_txt(filepath: str) -> list[dict]:
    entries = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()
        qa_pattern = r'[问Q][：:]\s*(.+?)\s*[答A][：:]\s*(.+?)(?=[问Q][：:]|\Z)'
        qa_matches = re.findall(qa_pattern, text, re.DOTALL)
        if qa_matches:
            for q, a in qa_matches:
                entries.append({"type": "qa", "title": q.strip()[:100], "content": a.strip()[:500]})
        else:
            paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
            if not paragraphs:
                paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
            for i, para in enumerate(paragraphs[:50]):
                entries.append({"type": "document", "title": f"段落{i+1}", "content": para[:1000]})
    except Exception as e:
        entries.append({"type": "error", "title": "文本解析失败", "content": str(e)})
    return entries


def parse_image(filepath: str) -> list[dict]:
    filename = Path(filepath).name
    if not DASHSCOPE_API_KEY:
        return [{"type": "image", "title": f"图片-{filename}", "content": f"[图片] {filename} - 未配置AI，需人工查看"}]
    try:
        content = _recognize_image(filepath)
        if content:
            print(f"  ✅ 图片AI识别成功: {filename}")
            return [{"type": "document", "title": f"图片内容-{filename}", "content": content}]
    except Exception as e:
        print(f"  ⚠️ 图片AI识别异常: {filename}, {e}")
    return [{"type": "image", "title": f"图片-{filename}", "content": f"[图片] {filename} - AI识别失败，需人工查看"}]


@retry(max_retries=2, base_delay=2.0)
def _recognize_image(filepath: str) -> str:
    """调用通义千问 VL 识别图片内容（带重试）"""
    with open(filepath, "rb") as f:
        img_data = base64.b64encode(f.read()).decode("utf-8")
    ext = Path(filepath).suffix.lower()
    mime = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
            ".gif": "image/gif", ".bmp": "image/bmp"}.get(ext, "image/png")
    resp = httpx.post(
        f"{DASHSCOPE_BASE_URL}/chat/completions",
        headers={"Authorization": f"Bearer {DASHSCOPE_API_KEY}", "Content-Type": "application/json"},
        json={"model": "qwen-vl-plus", "messages": [{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_data}"}},
            {"type": "text", "text": "请详细描述这张图片的内容。如果包含文字请提取所有文字。返回纯文本。"}
        ]}], "temperature": 0.3},
        timeout=60.0
    )
    if resp.status_code >= 500:
        raise Exception(f"DashScope VL 服务端错误: {resp.status_code}")
    if resp.status_code == 200:
        content = resp.json()["choices"][0]["message"]["content"]
        if content and content.strip():
            return content.strip()
    return ""


def parse_file(filepath: str, filename: str) -> list[dict]:
    """根据文件扩展名选择合适的解析器"""
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return parse_pdf(filepath)
    elif ext in (".doc", ".docx"):
        return parse_docx(filepath)
    elif ext in (".xls", ".xlsx"):
        return parse_excel(filepath)
    elif ext in (".txt", ".csv", ".md"):
        return parse_txt(filepath)
    elif ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp"):
        return parse_image(filepath)
    elif ext in (".zip", ".rar"):
        return [{"type": "archive", "title": f"压缩包-{filename}", "content": "[压缩包] 需解压后处理"}]
    else:
        return [{"type": "unknown", "title": f"未知格式-{filename}", "content": f"不支持: {ext}"}]
